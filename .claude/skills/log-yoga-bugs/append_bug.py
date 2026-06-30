#!/usr/bin/env python3
"""Append one or more user-identified bugs to bug-tracker.xlsx.

Usage:
    python3 append_bug.py bugs.json            # uses <repo-root>/bug-tracker.xlsx
    python3 append_bug.py bugs.json other.xlsx # explicit workbook path

bugs.json is a JSON list (or single object) with these keys per bug:
    original     (required) - the user's own description, in their words
    enhanced     (required) - refined / corrected understanding of the bug
    root_cause   (required) - the actual technical cause
    resolution   (required) - what was done to fix it (or "" if still open)
    status       (optional) - Open | In Progress | Fixed | Won't Fix | Not a Bug
                              (defaults to "Fixed" if a resolution is given,
                               otherwise "Open")
    reference    (optional) - commit / PR / files
    reported_by  (optional) - defaults to "Jessica"

Bug ID and Date Logged are assigned automatically. The script never edits or
removes existing rows.
"""
import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_XLSX = REPO_ROOT / "bug-tracker.xlsx"

# column order must match the headers in bug-tracker.xlsx
COLS = ["bug_id", "date", "reported_by", "original", "enhanced",
        "root_cause", "resolution", "status", "reference"]


def load_bugs(path):
    data = json.loads(Path(path).read_text())
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list) or not data:
        sys.exit("bugs JSON must be a non-empty object or list of objects")
    return data


def next_id_number(ws):
    nums = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = row[0]
        if isinstance(val, str) and val.upper().startswith("BUG-"):
            try:
                nums.append(int(val.split("-")[1]))
            except (IndexError, ValueError):
                pass
    return (max(nums) + 1) if nums else 1


def first_empty_row(ws):
    r = 2
    while ws.cell(row=r, column=1).value not in (None, ""):
        r += 1
    return r


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    bugs = load_bugs(sys.argv[1])
    xlsx = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_XLSX
    if not xlsx.exists():
        sys.exit(f"Workbook not found: {xlsx}")

    wb = load_workbook(xlsx)
    ws = wb["Bug Tracker"] if "Bug Tracker" in wb.sheetnames else wb.active

    body_font = Font(name="Arial", size=11)
    top_wrap = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D9D4CB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n = next_id_number(ws)
    row = first_empty_row(ws)
    logged = []
    today = date.today().isoformat()

    for bug in bugs:
        for req in ("original", "enhanced", "root_cause"):
            if not str(bug.get(req, "")).strip():
                sys.exit(f"bug missing required field: {req}")
        resolution = str(bug.get("resolution", "")).strip()
        status = str(bug.get("status", "")).strip()
        if not status:
            status = "Fixed" if resolution else "Open"
        bug_id = f"BUG-{n:03d}"
        values = {
            "bug_id": bug_id,
            "date": today,
            "reported_by": str(bug.get("reported_by", "Jessica")).strip() or "Jessica",
            "original": str(bug["original"]).strip(),
            "enhanced": str(bug["enhanced"]).strip(),
            "root_cause": str(bug["root_cause"]).strip(),
            "resolution": resolution,
            "status": status,
            "reference": str(bug.get("reference", "")).strip(),
        }
        for col_idx, key in enumerate(COLS, start=1):
            cell = ws.cell(row=row, column=col_idx, value=values[key])
            cell.font = body_font
            cell.alignment = top_wrap
            cell.border = border
        logged.append(f"{bug_id}: {values['original'][:60]}")
        n += 1
        row += 1

    wb.save(xlsx)
    print(f"Logged {len(logged)} bug(s) to {xlsx}:")
    for line in logged:
        print("  " + line)


if __name__ == "__main__":
    main()
